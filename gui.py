#!/usr/bin/env python3
"""
FPL Bill Extractor — Desktop GUI
=================================
Drag-and-drop a folder of FPL bills (or use the Browse button), pick an
output location, and click Extract. Results land in a formatted .xlsx.

Works standalone or bundled as a macOS .app via PyInstaller.
"""
import os
import sys
import glob
import threading
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# When bundled by PyInstaller, modules are inside the temp dir.
if getattr(sys, "frozen", False):
    sys.path.insert(0, sys._MEIPASS)

import fpl_parser
import rates

BILL_EXTS = (".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp")

# ── Column spec (same as CLI version) ────────────────────────────────────────
COLUMNS = [
    ("source_file", "Source File"),
    ("extract_method", "Extract Method"),
    ("account_number", "Account #"),
    ("meter_number", "Meter"),
    ("rate_schedule", "Rate"),
    ("statement_date", "Stmt Date"),
    ("service_period_start", "Period Start"),
    ("service_period_end", "Period End"),
    ("days_in_period", "Days"),
    ("kwh_total", "kWh Total"),
    ("kwh_on_peak", "kWh On-Pk"),
    ("kwh_off_peak", "kWh Off-Pk"),
    ("demand_kw_billed", "Billed kW"),
    ("demand_kw_actual", "Actual kW"),
    ("total_amount_due", "Amount Due"),
    ("min_demand_kw", "Min kW"),
    ("load_factor", "Load Factor"),
    ("breakeven_peak_kw", "Breakeven kW"),
    ("est_gs1", "Est GS-1"),
    ("est_gsd1", "Est GSD-1"),
    ("recommended_rate", "Rec. Rate"),
    ("flag", "Flag"),
]


def extract_one(path):
    """Extract fields from a single bill and run the rate screen."""
    rec = {c: None for c, _ in COLUMNS}
    rec["source_file"] = os.path.basename(path)
    text, method = fpl_parser.get_text(path)
    fields = fpl_parser.parse_fields(text)
    rec["extract_method"] = method
    for k, v in fields.items():
        if k in rec:
            rec[k] = v
    peak = rec.get("demand_kw_billed") or rec.get("demand_kw_actual")
    scr = rates.screen(rec.get("rate_schedule"), rec.get("kwh_total"),
                       peak_kw=peak, days=rec.get("days_in_period"))
    for k in ("min_demand_kw", "load_factor", "breakeven_peak_kw",
              "est_gs1", "est_gsd1", "recommended_rate"):
        rec[k] = scr[k]
    if scr["flag"]:
        rec["flag"] = scr["flag"]
    return rec


def write_xlsx(records, out_path):
    """Write extracted records to a formatted spreadsheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Bills"
    ws.sheet_view.showGridLines = False
    arial = "Arial"
    hdr_font = Font(name=arial, bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c, (_, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    money = {"total_amount_due", "est_gs1", "est_gsd1"}
    num0 = {"kwh_total", "kwh_on_peak", "kwh_off_peak"}
    num1 = {"demand_kw_billed", "demand_kw_actual", "min_demand_kw",
            "breakeven_peak_kw"}
    for r, rec in enumerate(records, 2):
        for c, (key, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(r, c, rec.get(key))
            cell.font = Font(name=arial, size=10)
            cell.border = border
            if key in money:
                cell.number_format = "$#,##0.00"
            elif key in num0:
                cell.number_format = "#,##0"
            elif key in num1:
                cell.number_format = "#,##0.0"
            elif key == "load_factor":
                cell.number_format = "0.0%"
            if (key == "flag" and isinstance(rec.get("flag"), str)
                    and "MISCLASS" in rec["flag"]):
                for cc in range(1, len(COLUMNS) + 1):
                    ws.cell(r, cc).fill = PatternFill("solid", fgColor="FCE4E4")

    widths = [22, 10, 12, 10, 8, 11, 11, 11, 6, 10, 9, 9, 9, 9, 11,
              9, 8, 10, 9, 9, 14, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 28
    wb.save(out_path)


# ── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FPL Bill Extractor")
        self.geometry("680x520")
        self.configure(bg="#1e2530")
        self.resizable(False, False)

        self.bill_files = []
        self.folder_path = tk.StringVar(value="")
        self.out_path = tk.StringVar(
            value=os.path.join(os.path.expanduser("~/Desktop"),
                               "fpl_bills.xlsx"))

        self._build_ui()
        self._try_enable_dnd()

    # ── UI layout ────────────────────────────────────────────────────────
    def _build_ui(self):
        bg = "#1e2530"
        fg = "#e0e0e0"
        accent = "#3b82f6"
        entry_bg = "#2a3240"
        btn_bg = "#334155"

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TLabel", background=bg, foreground=fg,
                        font=("Helvetica", 12))
        style.configure("Title.TLabel", background=bg, foreground="#ffffff",
                        font=("Helvetica", 18, "bold"))
        style.configure("Sub.TLabel", background=bg, foreground="#94a3b8",
                        font=("Helvetica", 10))
        style.configure("TButton", background=btn_bg, foreground=fg,
                        font=("Helvetica", 11), padding=(12, 6))
        style.map("TButton",
                  background=[("active", accent), ("disabled", "#1e2530")])
        style.configure("Accent.TButton", background=accent,
                        foreground="#ffffff", font=("Helvetica", 13, "bold"),
                        padding=(16, 10))
        style.map("Accent.TButton",
                  background=[("active", "#2563eb"), ("disabled", "#475569")])
        style.configure("Horizontal.TProgressbar", troughcolor=entry_bg,
                        background=accent, thickness=8)

        # Title
        ttk.Label(self, text="FPL Bill Extractor",
                  style="Title.TLabel").pack(pady=(20, 2))
        ttk.Label(self, text="Drag a folder of bills below, or browse",
                  style="Sub.TLabel").pack()

        # Drop zone
        self.drop_frame = tk.Frame(self, bg="#2a3240", highlightbackground="#475569",
                                   highlightthickness=2, cursor="hand2")
        self.drop_frame.pack(padx=30, pady=(16, 10), fill="x", ipady=28)

        self.drop_label = tk.Label(
            self.drop_frame, text="📂  Drop folder here  📂",
            bg="#2a3240", fg="#94a3b8", font=("Helvetica", 14))
        self.drop_label.pack(expand=True)

        self.drop_frame.bind("<Button-1>", lambda e: self._browse_folder())
        self.drop_label.bind("<Button-1>", lambda e: self._browse_folder())

        # Folder path display
        path_frame = tk.Frame(self, bg=bg)
        path_frame.pack(padx=30, fill="x")
        ttk.Label(path_frame, text="Bills folder:").pack(side="left")
        self.path_display = tk.Label(path_frame, textvariable=self.folder_path,
                                     bg=bg, fg=accent, font=("Helvetica", 10),
                                     anchor="w")
        self.path_display.pack(side="left", padx=(6, 0), fill="x", expand=True)

        # Output row
        out_frame = tk.Frame(self, bg=bg)
        out_frame.pack(padx=30, pady=(10, 0), fill="x")
        ttk.Label(out_frame, text="Save to:").pack(side="left")
        tk.Entry(out_frame, textvariable=self.out_path, bg=entry_bg, fg=fg,
                 insertbackground=fg, font=("Helvetica", 10), bd=0,
                 highlightthickness=1, highlightcolor=accent,
                 highlightbackground="#475569").pack(
                     side="left", padx=(6, 6), fill="x", expand=True)
        ttk.Button(out_frame, text="…",
                   command=self._browse_output).pack(side="right")

        # Progress
        self.progress = ttk.Progressbar(self, mode="determinate",
                                        style="Horizontal.TProgressbar")
        self.progress.pack(padx=30, pady=(16, 4), fill="x")

        self.status = tk.Label(self, text="Ready", bg=bg, fg="#94a3b8",
                               font=("Helvetica", 10), anchor="w")
        self.status.pack(padx=32, fill="x")

        # Extract button
        self.extract_btn = ttk.Button(self, text="Extract Bills",
                                      style="Accent.TButton",
                                      command=self._run_extract)
        self.extract_btn.pack(pady=(14, 20))

    # ── Drag-and-drop (optional; works without it) ───────────────────────
    def _try_enable_dnd(self):
        """Try to enable native drag-and-drop via tkinterdnd2."""
        try:
            import tkinterdnd2
            # Re-init as a DnD-capable window
            self.drop_frame.drop_target_register(tkinterdnd2.DND_FILES)
            self.drop_frame.dnd_bind("<<Drop>>", self._on_drop)
            self.drop_frame.dnd_bind("<<DragEnter>>",
                                     lambda e: self.drop_frame.configure(
                                         highlightbackground="#3b82f6"))
            self.drop_frame.dnd_bind("<<DragLeave>>",
                                     lambda e: self.drop_frame.configure(
                                         highlightbackground="#475569"))
        except Exception:
            # tkinterdnd2 not available; browse button still works fine.
            pass

    def _on_drop(self, event):
        path = event.data.strip("{}")  # macOS wraps paths in braces
        if os.path.isdir(path):
            self._set_folder(path)
        elif os.path.isfile(path):
            self._set_folder(os.path.dirname(path))

    # ── Folder / output selection ────────────────────────────────────────
    def _browse_folder(self):
        path = filedialog.askdirectory(title="Select folder of FPL bills")
        if path:
            self._set_folder(path)

    def _set_folder(self, path):
        self.folder_path.set(path)
        self.bill_files = sorted(
            f for f in glob.glob(os.path.join(path, "*"))
            if f.lower().endswith(BILL_EXTS))
        n = len(self.bill_files)
        self.drop_label.configure(
            text=f"📂  {os.path.basename(path)}  —  {n} bill{'s' if n != 1 else ''} found",
            fg="#e0e0e0" if n else "#ef4444")
        self.drop_frame.configure(highlightbackground="#475569")

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Save spreadsheet as",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="fpl_bills.xlsx")
        if path:
            self.out_path.set(path)

    # ── Extraction (threaded so the GUI stays responsive) ────────────────
    def _run_extract(self):
        if not self.bill_files:
            messagebox.showwarning("No bills",
                                   "Select a folder with PDF or image bills first.")
            return
        out = self.out_path.get().strip()
        if not out:
            messagebox.showwarning("No output", "Choose an output file path.")
            return

        self.extract_btn.configure(state="disabled")
        self.progress["value"] = 0
        self.progress["maximum"] = len(self.bill_files)
        threading.Thread(target=self._extract_thread, args=(out,),
                         daemon=True).start()

    def _extract_thread(self, out_path):
        records = []
        errors = 0
        for i, path in enumerate(self.bill_files):
            name = os.path.basename(path)
            self._update_status(f"[{i+1}/{len(self.bill_files)}]  {name}")
            try:
                rec = extract_one(path)
            except Exception:
                rec = {c: None for c, _ in COLUMNS}
                rec["source_file"] = name
                rec["flag"] = ("EXTRACT ERROR: "
                               + traceback.format_exc(limit=1).strip()
                               .replace("\n", " "))
                errors += 1
            records.append(rec)
            self._update_progress(i + 1)

        try:
            write_xlsx(records, out_path)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror(
                "Write error", f"Could not write spreadsheet:\n{e}"))
            self.after(0, lambda: self.extract_btn.configure(state="normal"))
            return

        ok = sum(1 for r in records if r.get("kwh_total"))
        msg = f"Done — {len(records)} bills, {ok} with consumption parsed"
        if errors:
            msg += f", {errors} errors"
        self._update_status(msg)
        self.after(0, lambda: self.extract_btn.configure(state="normal"))
        self.after(0, lambda: messagebox.showinfo(
            "Complete",
            f"{msg}.\n\nSaved to:\n{out_path}"))

    def _update_status(self, text):
        self.after(0, lambda: self.status.configure(text=text))

    def _update_progress(self, value):
        self.after(0, lambda: self.progress.configure(value=value))


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
