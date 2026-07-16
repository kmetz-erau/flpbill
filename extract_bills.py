#!/usr/bin/env python3
"""
FPL bill extraction tool
========================
Point it at a folder of FPL bills (PDF or image, digital or scanned); it
extracts the fields that matter across every rate class, computes load factor
and the GS-1 <-> GSD-1 rate-fit flag for each account, and writes one
formatted spreadsheet.

Digital PDFs are read from their text layer. Bills with no text layer (true
scans) fall back to tesseract OCR automatically -- per file, no sorting
needed. The `Extract Method` column in the output records which path each bill
took ('text-layer' or 'ocr'), so OCR'd rows are easy to spot-check.

Usage
-----
    python extract_bills.py ./bills -o portfolio.xlsx
    python extract_bills.py ./bills --force-ocr        # force OCR on every PDF
"""
import argparse
import glob
import os
import sys
import traceback

import rates
import fpl_parser

BILL_EXTS = (".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp")

# Column order for the output sheet.
COLUMNS = [
    ("source_file", "Source File"),
    ("extract_method", "Extract\nMethod"),
    ("account_number", "Account #"),
    ("meter_number", "Meter"),
    ("service_address", "Service Address"),
    ("rate_schedule", "Rate"),
    ("statement_date", "Stmt Date"),
    ("service_period_start", "Period Start"),
    ("service_period_end", "Period End"),
    ("days_in_period", "Days"),
    ("kwh_total", "kWh Total"),
    ("kwh_on_peak", "kWh On-Pk"),
    ("kwh_off_peak", "kWh Off-Pk"),
    ("demand_kw_billed", "Billed kW"),
    ("demand_kw_actual", "Actual kW"),
    ("total_amount_due", "Amount Due"),
    ("min_demand_kw", "Min kW\n(kWh/hrs)"),
    ("load_factor", "Load\nFactor"),
    ("breakeven_peak_kw", "Breakeven\nkW"),
    ("est_gs1", "Est GS-1"),
    ("est_gsd1", "Est GSD-1"),
    ("recommended_rate", "Rec. Rate"),
    ("flag", "Flag"),
]

# Fields the parser returns that are not in COLUMNS but feed the screen.
_EXTRA_PARSE_FIELDS = ("service_address", "demand_kw_on_peak")


def extract_one(path, force_ocr):
    rec = {c: None for c, _ in COLUMNS}
    rec["source_file"] = os.path.basename(path)

    text, method = fpl_parser.get_text(path, force_ocr=force_ocr)
    fields = fpl_parser.parse_fields(text)
    rec["extract_method"] = method
    for k, v in fields.items():
        if k in rec:
            rec[k] = v

    # derive metrics + rate-fit flag
    peak = rec.get("demand_kw_billed") or rec.get("demand_kw_actual")
    scr = rates.screen(rec.get("rate_schedule"), rec.get("kwh_total"),
                       peak_kw=peak, days=rec.get("days_in_period"))
    for k in ("min_demand_kw", "load_factor", "breakeven_peak_kw",
              "est_gs1", "est_gsd1", "recommended_rate"):
        rec[k] = scr[k]
    if scr["flag"]:
        rec["flag"] = scr["flag"]
    return rec


def write_xlsx(records, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Bills"
    ws.sheet_view.showGridLines = False
    arial = "Arial"
    hdr_font = Font(name=arial, bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c, (_, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, label)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center

    money = {"total_amount_due", "est_gs1", "est_gsd1"}
    num0 = {"kwh_total", "kwh_on_peak", "kwh_off_peak"}
    num1 = {"demand_kw_billed", "demand_kw_actual", "min_demand_kw",
            "breakeven_peak_kw"}
    for r, rec in enumerate(records, 2):
        for c, (key, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(r, c, rec.get(key))
            cell.font = Font(name=arial, size=10)
            cell.border = border
            if key in money:
                cell.number_format = '$#,##0.00'
            elif key in num0:
                cell.number_format = '#,##0'
            elif key in num1:
                cell.number_format = '#,##0.0'
            elif key == "load_factor":
                cell.number_format = '0.0%'
            if key == "flag" and isinstance(rec.get("flag"), str) and "MISCLASS" in rec["flag"]:
                for cc in range(1, len(COLUMNS) + 1):
                    ws.cell(r, cc).fill = PatternFill("solid", fgColor="FCE4E4")

    widths = [22, 10, 12, 10, 24, 8, 11, 11, 11, 6, 10, 9, 9, 9, 9, 11,
              9, 8, 10, 9, 9, 14, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 28
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Extract FPL bills into a spreadsheet.")
    ap.add_argument("folder", help="folder of bill PDFs / images")
    ap.add_argument("-o", "--out", default="fpl_bills.xlsx", help="output .xlsx")
    ap.add_argument("--force-ocr", action="store_true", help="OCR every PDF (all scans)")
    args = ap.parse_args()

    files = sorted(f for f in glob.glob(os.path.join(args.folder, "*"))
                   if f.lower().endswith(BILL_EXTS))
    if not files:
        sys.exit(f"No bill files found in {args.folder}")

    records = []
    for i, path in enumerate(files, 1):
        print(f"[{i}/{len(files)}] {os.path.basename(path)} ...", end=" ", flush=True)
        try:
            rec = extract_one(path, args.force_ocr)
            print(f"{rec.get('rate_schedule') or '?'} | "
                  f"{rec.get('kwh_total') or '?'} kWh | {rec.get('flag') or ''}")
        except Exception:
            rec = {c: None for c, _ in COLUMNS}
            rec["source_file"] = os.path.basename(path)
            rec["flag"] = "EXTRACT ERROR: " + traceback.format_exc(limit=1).strip().replace("\n", " ")
            print("ERROR")
        records.append(rec)

    write_xlsx(records, args.out)
    ok = sum(1 for r in records if r.get("kwh_total"))
    print(f"\nWrote {args.out}: {len(records)} bills, {ok} with consumption parsed.")


if __name__ == "__main__":
    main()
